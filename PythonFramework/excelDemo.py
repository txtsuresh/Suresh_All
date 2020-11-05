import openpyxl

excel=openpyxl.load_workbook("C:\\Users\\Suresh Kumar\\Desktop\\Automation_Datadriven.xlsx")
sheet=excel.active #Navigates to excel sheet
cell=sheet.cell(row=1, column=2) #Navigates to the cell in the sheet
print(cell.value) #REad cell value
print(sheet['A5'].value) #REad cell value

sheet.cell(row=1, column=2).value="suresh" #Write cell value
print(sheet.cell(row=1, column=2).value)
print(sheet.max_column, sheet.max_row) #prints max. column & max. rows

for i in range(1,sheet.max_row+1): # Prints all values in the excel sheet
    for j in range(1,sheet.max_column+1):
        print(sheet.cell(row=i,column=j).value)


for i in range(1,sheet.max_row+1): # Prints all values from row#2 in the excel sheet
    if sheet.cell(row=i,column=1).value=="FUSE":
        for j in range(1,sheet.max_column+1):
            print(sheet.cell(row=i,column=j).value)

#Add to dictionary
Dict={}
for i in range(1,sheet.max_row+1): # Prints all values from row#2 in the excel sheet
    if sheet.cell(row=i,column=1).value=="FUSE":
        for j in range(1,sheet.max_column+1):
            print(sheet.cell(row=i,column=j).value)
            Dict[sheet.cell(row=1,column=j).value]=sheet.cell(row=i,column=j).value
print(Dict)