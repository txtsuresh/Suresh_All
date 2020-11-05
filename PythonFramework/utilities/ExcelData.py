import openpyxl

class ExcelData:

    @staticmethod
    def getTestData(test_name):
        Dict = {}
        excel=openpyxl.load_workbook("C:\\Users\\Suresh Kumar\\Desktop\\Automation_Datadriven.xlsx")
        sheet=excel.active #Navigates to excel sheet


        for i in range(1,sheet.max_row+1): # Prints all values from row#2 in the excel sheet
            if sheet.cell(row=i,column=1).value==test_name:
                for j in range(2,sheet.max_column+1):
                    Dict[sheet.cell(row=1,column=j).value]=sheet.cell(row=i,column=j).value
        return [Dict]